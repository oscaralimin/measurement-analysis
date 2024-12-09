import logging
from pathlib import Path
import openpyxl
from typing import Dict, Any

from .validators import ConfigValidator

class ConfigHandler:
    """
    Handles loading and managing configuration from Excel files.
    """
    def __init__(self, config_file: Path):
        self.logger = logging.getLogger(__name__)
        self.config_file = config_file
        self.validator = ConfigValidator()
        self.config = {}
        self.load_config()

    def load_config(self):
        """
        Loads configuration from Excel file.
        Validates all channel configurations.
        """
        try:
            self.logger.info(f"Loading configuration from {self.config_file}")
            wb = openpyxl.load_workbook(self.config_file)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1][1:]]
            
            # Process each channel row
            for row in ws.iter_rows(min_row=2):
                if not row[0].value:  # Skip empty rows
                    continue
                    
                channel_name = row[0].value
                # Create dictionary of parameters for this channel
                config_data = {
                    headers[i]: row[i+1].value if row[i+1].value is not None else ""
                    for i in range(len(headers))
                }
                
                # Validate and store configuration
                if self.validator.validate_channel_config(channel_name, config_data):
                    self.config[channel_name] = config_data
                else:
                    self.logger.warning(f"Skipping invalid configuration for channel {channel_name}")
                    
        except Exception as e:
            self.logger.error(f"Failed to load configuration: {str(e)}")
            raise

    def get_channel_config(self, channel_name: str) -> Dict[str, Any]:
        """
        Retrieves configuration for a specific channel.
        
        Args:
            channel_name: Name of the channel
            
        Returns:
            Dictionary containing channel configuration
        """
        if channel_name not in self.config:
            raise KeyError(f"No configuration found for channel {channel_name}")
        return self.config[channel_name]

    def update_channel_config(self, channel_name: str, new_config: Dict[str, Any]) -> bool:
        """
        Updates configuration for a specific channel.
        
        Args:
            channel_name: Name of the channel
            new_config: New configuration dictionary
            
        Returns:
            bool: True if update successful, False otherwise
        """
        try:
            if self.validator.validate_channel_config(channel_name, new_config):
                self.config[channel_name] = new_config
                return True
            return False
        except Exception as e:
            self.logger.error(f"Failed to update config for {channel_name}: {str(e)}")
            return False