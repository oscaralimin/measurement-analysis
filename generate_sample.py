# generate_samples.py
import numpy as np
import asammdf
from asammdf import MDF, Signal
import openpyxl
from datetime import datetime

def create_sample_excel():
    """Creates sample Excel configuration file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define headers
    headers = [
        "Channel Name",
        "Sollwertkanal",
        "Toleranz statisch",
        "Skalierung",
        "back2backID",
        "back2backIDPosition",
        "Sollwertkanalskalierung",
        "Sollwert statisch",
        "Testflagchannel",
        "startTime",
        "endTime",
        "Unit"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample data
    sample_channels = [
        {
            "Channel Name": "Temperature_Engine",
            "Sollwertkanal": "Temp_Setpoint",
            "Toleranz statisch": "5.0",
            "Skalierung": "1.0",
            "back2backID": "ENG",
            "back2backIDPosition": "0",
            "Sollwertkanalskalierung": "1.0",
            "Sollwert statisch": "",
            "Testflagchannel": "",
            "startTime": "",
            "endTime": "",
            "Unit": "°C"
        },
        {
            "Channel Name": "Pressure_System",
            "Sollwertkanal": "",
            "Toleranz statisch": "10.0",
            "Skalierung": "1.0",
            "back2backID": "SYS",
            "back2backIDPosition": "0",
            "Sollwertkanalskalierung": "",
            "Sollwert statisch": "100.0",
            "Testflagchannel": "",
            "startTime": "",
            "endTime": "",
            "Unit": "bar"
        }
    ]
    
    # Write data
    for row, channel in enumerate(sample_channels, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=channel.get(header, ""))
    
    # Save file
    wb.save("config.xlsx")
    print("Created config.xlsx")

def create_sample_mf4():
    timestamps = np.arange(0, 10, 0.01)
    
    mdf = MDF()
    
    # Temperature Engine
    temp_engine = 90 + 2*np.sin(timestamps) + np.random.normal(0, 0.5, len(timestamps))
    mdf.append(
        Signal(samples=temp_engine, timestamps=timestamps, name='Temperature_Engine', unit='°C')
    )
    
    # Temperature Setpoint
    temp_setpoint = np.full_like(timestamps, 90)
    mdf.append(
        Signal(samples=temp_setpoint, timestamps=timestamps, name='Temp_Setpoint', unit='°C')
    )
    
    # Pressure System
    pressure = 100 + np.random.normal(0, 1, len(timestamps))
    pressure[300:310] += 15
    mdf.append(
        Signal(samples=pressure, timestamps=timestamps, name='Pressure_System', unit='bar')
    )
    
    mdf.save('sample_measurement.mf4', overwrite=True)

if __name__ == "__main__":
    create_sample_excel()
    create_sample_mf4()